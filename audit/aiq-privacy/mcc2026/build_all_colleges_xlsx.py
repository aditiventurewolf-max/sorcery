import json, re, glob, os, collections, urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE=os.path.dirname(os.path.abspath(__file__)); RC=os.path.join(HERE,'rc')
full=json.load(open(os.path.join(HERE,'full135.json')))
hostmap={r['host']:r for r in json.load(open(os.path.join(RC,'hostmap.json')))}

# ---- calendars -------------------------------------------------------------
cal={}
for f in glob.glob(os.path.join(RC,'cal_*.log'))+glob.glob(os.path.join(RC,'p_cal_*.log')):
    host=None
    for line in open(f, errors='ignore'):
        m=re.match(r'^([a-z0-9][a-z0-9._-]+)\s{2,}(HAS CALENDAR|no academic calendar found)', line)
        if m:
            host=m.group(1); cal[host]={'has': m.group(2)=='HAS CALENDAR', 'ev':''}
        elif host and cal.get(host,{}).get('has') and line.strip().startswith('http') and not cal[host]['ev']:
            cal[host]['ev']=line.strip()

# ---- fees ------------------------------------------------------------------
fees=collections.defaultdict(list)
for f in (glob.glob(os.path.join(RC,'fee_*.log'))+glob.glob(os.path.join(RC,'p_fee_*.log'))
          +glob.glob(os.path.join(RC,'w_fee_*.log'))):
    host=None
    for line in open(f, errors='ignore'):
        m=re.match(r'^--- (\S+)', line)
        if m: host=m.group(1); continue
        m=re.match(r'\s+(TUITION|TOTAL/DD)\s+Rs\s+([\d,]+)\s+(\S.*)$', line)
        if m is None:
            m2=re.match(r'^### (\S+)', line)
            if m2: host=m2.group(1)
        if m and host:
            fees[host].append((m.group(1), int(m.group(2).replace(',','')), m.group(3).strip()))

# state-level figures CONFIRMED from >=2 colleges' own 2026-27 documents
STATE_CONFIRMED={'Maharashtra':(167300,'Confirmed 2026-27 from own admission brochures of GMC Bhandara, Ratnagiri, '
                                       'Baramati, Nandurbar, Yavatmal, VMGMC Miraj and MGIMS Sevagram (eight colleges, all ₹1,67,300)')}

def best_tuition(host, state):
    """Prefer an MBBS/AIQ 2026-27 figure from this college's own documents."""
    cands=[c for c in fees.get(host,[]) if c[0]=='TUITION']
    mbbs=[c for c in cands
          if re.search(r'fee|brochure|brouchure|prospect|instruction|guideline|structure|proforma|manual|information', c[2], re.I)
          and not re.search(r'\bpg\b|post.?grad|nursing|msc|m\.sc|paramedic|dnb|naac', c[2], re.I)
          and not re.search(r'admission[-_ ]?list|student[-_ ]?list|batch|result|retention|\bdata\b', c[2], re.I)]
    pool=mbbs or []
    y26=[c for c in pool if re.search(r'2026', c[2])]
    pick=(y26 or pool)
    if pick:
        v=max(x[1] for x in pick)
        src=[x[2] for x in pick if x[1]==v][0]
        clash=[st for st,(sv,_) in STATE_CONFIRMED.items() if sv==v and st!=state]
        if clash:
            return None, ('NOT CONFIRMED — the only figure found (Rs %s) equals the confirmed %s state figure '
                          'and is not credible for %s; treated as a false match' % (f'{v:,}', clash[0], state)), 'none'
        return v, 'College document: '+src, 'college'
    if state in STATE_CONFIRMED:
        v,note=STATE_CONFIRMED[state]
        return v, note+' — NOT confirmed at this college (it publishes no fee document)', 'state'
    return None, 'NOT CONFIRMED — no fee document found on this college site', 'none'

# ---- 2026 R1 ranks ---------------------------------------------------------
rows=json.load(open(os.path.join(HERE,'rows.json')))
aiq=[r for r in rows if r['inst'].startswith('All India ') and r['course']=='MBBS'
     and r['ccat']=='General' and r['acat'].strip()=='Open']
by=collections.defaultdict(list)
for r in aiq: by[r['inst'][len('All India '):].strip()].append(r['rank'])
STOP={'government','govt','goverment','medical','college','hospital','institute','of','and','the',
      'sciences','science','dr','memorial','research','centre','center','general','district','shri',
      'smt','late','rural','autonomous','state','society','allopathic','rajkiya','instt','inst',
      'previously','known','as','formerly','renamed','hospitals','health','super','facility'}
def toks(s):
    return {w for w in re.sub(r'[^A-Za-z ]',' ',(s or '')).lower().split()
            if w not in STOP and len(w)>3}

# inverse document frequency over the MCC institute strings
_df=collections.Counter()
for k in by: 
    for w in toks(k): _df[w]+=1

# explicit aliases where the MCC string and the CSV name share no distinctive token
ALIAS={
 'Sri Venkateswara Medical College, Tirupati':'S V Medical College, Tirupati',
 'Sheikh Bhikhari Medical College & Hospital, Hazaribag':'Hazaribagh Medical College',
 'Chhattisgarh Institute Of Medical Sciences, Bilaspur':'Chhattisgarh Institute of Medical Sciences, Bilaspur',
 'Dr. Vaishampayam Memorial M.C., Sholapur':'Dr Vaishampayan Memorial Medical College, Solapur',
 'Neigrihms, Shillong':'North Eastern Indira Gandhi Regional Instt. of Health and Medical Sciences, Shillong',
 'Gulbarga Institute Of Medical Sciences, Gulbarga':'Gulbarga Institute of Medical Sciences',
}
def close26(college):
    college = ALIAS.get(college, college)
    t=toks(college)
    if not t: return None
    scored=[]
    for k,v in by.items():
        kt=toks(k)
        common=t & kt
        if not common: continue
        # weight rare tokens far above common ones
        score=sum(1.0/_df[w] for w in common if _df[w])
        rare=min((_df[w] for w in common), default=99)
        scored.append((score, rare, max(v), k))
    if not scored: return None
    scored.sort(reverse=True)
    best=scored[0]
    # need a genuinely distinctive shared token, and a clear win over the runner-up
    if best[1] > 3: return None
    if len(scored) > 1 and scored[1][0] >= best[0]: return None
    return best[2]

# ---- session findings (documents actually found) ---------------------------
SESSION={
 'GMC Shivpuri':'files/BATCH 2023-24.pdf and BATCH 2024-25.pdf — "Name of Student" + NEET percentile per row; PG merit list with names+percentiles',
 'Lt. L A M Govt. Medical College, Raigarh':'storage/grievances.csv — 4 records marked Confidential=Yes with name, email, phone, complaint text; storage/feedback.csv (Name,Email,Message)',
 'Sh Vasant Rao Naik Govt.M.C., Yavatmal':'studentsection/U2015,U2016,U2019,U2020.pdf — "Name of the Student / M-F / Date of admission / Category"; plus 7 MBBS exam-result PDFs with names and seat numbers',
 'Rims, Ongole':'UG-STUDENTS-LIST-2022-BATCH.pdf — name + roll number + 10-digit mobile; monthly PG stipend lists; 3 NMC00*.pdf; 13 hits total',
 'Govt Medical College Baramati':'brcp/studet_list_mbbs/ — batch 2020-21, 2021-22, 2022-23, each with a student-name header',
 'Medinirai Medical College (Previously Known As Palamau Medical College), Daltonganj':'Student-List-25-26.pdf (CURRENT intake), MBBS-BATCH-2023.pdf (89 rows), Student-List-2019.pdf (86 rows), MBBS-BATCH-2022.pdf, List-of-Candidate-2024-25.pdf',
 'Government Medical College And General Hospital, Jalgaon':'assets/pdf/admission/ug/UG Data-2025-26.pdf — current 2025-26 intake, 2.2 MB scan',
 'College Of Medicine And Jnm Hospital, Kalyani':'assets/pdf/<hash>.pdf — hash-named roster PDF (no meaningful link text)',
 'Sagar Dutta Medical College & Hospital, Kolkata':'class lists + full batch lists (only visible at long timeouts — slow server)',
}
CLEAN={ # verified clean this session, with the depth
 'Government Medical College, Latur':'NOTHING FOUND. Site is an 11-page brochure with zero documents of any kind (17 URLs = full coverage).',
 'Government Medical College, Akola':'NOTHING FOUND. Documents served from api.gmcakola.in; store empty across Type 0-12 x IsWeb true/false.',
 'Swami Ramanand Tirth Rural M.C, Ambajogai':'388 docs opened. Only MUHS mandatory-disclosure filings (faculty/beds/dean); the 2025-26 one is an unfilled template.',
 'Govt. Medical College And Hospital, Chandrapur':'186 docs opened. Only a Marathi circular and a library journal-subscription list.',
 'Government Medical College, Seoni':'107 docs opened. Only faculty/resident AEBAS attendance (professor x51, resident x28).',
 'Government Medical College, Ratnagiri':'75 docs opened. Only admission-guideline brochures with blank signature/father fields.',
 'Government Medical College, Bhandara':'413 fixed paths + 25 docs. Only admission instruction brochures (blank form fields).',
 'Shaheed Nirmal Mahto Medical College & Hospital, Dhanbad':'FULL BATTERY: 413 fixed paths, 89 URLs, no open dirs, 31 docs opened, 0 hits. Two scanned PDFs OCR-read: subject/exam schedules, no names.',
 'Government Medical College, Baramulla':'NOTHING FOUND. 93 KB static site, zero .pdf references in markup, 124 URLs.',
}

wb=openpyxl.Workbook(); ws=wb.active; ws.title='All colleges - filters'
HDR=Font(bold=True,color='FFFFFF',size=10); HFIL=PatternFill('solid',fgColor='1F4E79')
WRAP=Alignment(wrap_text=True,vertical='top'); CEN=Alignment(horizontal='center',vertical='center')
THIN=Border(*[Side(style='thin',color='D9D9D9')]*4)
GOOD=PatternFill('solid',fgColor='C6EFCE'); BAD=PatternFill('solid',fgColor='FFC7CE')
WARN=PatternFill('solid',fgColor='FFEB9C'); GREY=PatternFill('solid',fgColor='F2F2F2')

COLS=[('College',46),('State',18),('2026 R1 close',12),('2025 R1 close',12),
      ('Academic calendar?',14),('Calendar evidence',44),
      ('Confirmed tuition Rs/yr',14),('Under Rs 40,000?',13),('Fee basis',66),
      ('Documents found on the site',86),('Evidence URL',52),('Site',30)]
for j,(h,w) in enumerate(COLS,start=1):
    c=ws.cell(row=1,column=j,value=h); c.font=HDR; c.fill=HFIL
    c.alignment=Alignment(wrap_text=True,horizontal='center',vertical='center')
    ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[1].height=42

def hostof(r):
    u=(r.get('url') or '').split(' ')[0].strip()
    return urllib.parse.urlparse(u).netloc if u.startswith('http') else ''

out=[]
for r in full:
    host=hostof(r)
    c=cal.get(host)
    calv = 'site not reachable' if not host else ('Yes' if (c and c['has']) else ('No' if c else 'not checked'))
    calev = (c or {}).get('ev','') if c else ''
    tv, basis, kind = best_tuition(host, r['state']) if host else (None,'Site not reachable from this environment — no fee document to read','none')
    if not host and r['state'] in STATE_CONFIRMED:
        tv=STATE_CONFIRMED[r['state']][0]
        basis=STATE_CONFIRMED[r['state']][1]+' — NOT confirmed at this college (site unreachable)'; kind='state'
    under = 'Unknown' if tv is None else ('Yes' if tv < 40000 else 'No')
    docs = SESSION.get(r['college']) or CLEAN.get(r['college']) or (r.get('sessions') or '').strip()
    if not docs:
        docs = ('No documents found on this site' if r['verdict']=='PASS'
                else 'Not audited — site unreachable from this environment' if r['verdict']=='PROVISIONAL'
                else '')
    out.append({'college':r['college'],'state':r['state'],'c26':close26(r['college']),
                'c25':r['close25'],'cal':calv,'calev':calev,'tv':tv,'under':under,
                'basis':basis,'docs':docs,'ev':r.get('disc_url') or r.get('scribd_url') or '','host':host})

out.sort(key=lambda x:(x['c26'] is None, x['c26'] or 0))
for i,o in enumerate(out,start=2):
    vals=[o['college'],o['state'],o['c26'],o['c25'],o['cal'],o['calev'],o['tv'],o['under'],
          o['basis'],o['docs'],o['ev'],o['host']]
    for j,v in enumerate(vals,start=1):
        cell=ws.cell(row=i,column=j,value=v if v not in (None,'') else ('' if j not in (3,7) else 'n/a'))
        cell.border=THIN
        cell.alignment=WRAP if j in (6,9,10,11) else CEN
    ws.cell(row=i,column=7).number_format='#,##0'
    ws.cell(row=i,column=5).fill = GOOD if o['cal']=='No' else (BAD if o['cal']=='Yes' else GREY)
    ws.cell(row=i,column=8).fill = GOOD if o['under']=='Yes' else (BAD if o['under']=='No' else GREY)
    if o['docs'].startswith(('NOTHING FOUND','No documents','FULL BATTERY')): ws.cell(row=i,column=10).fill=GOOD
    elif o['docs'].startswith('Not audited'): ws.cell(row=i,column=10).fill=GREY
    elif o['docs']: ws.cell(row=i,column=10).fill=BAD
ws.auto_filter.ref=f'A1:L{len(out)+1}'
ws.freeze_panes='C2'

# notes sheet
ns=wb.create_sheet('Notes on the fee column')
N=[('Rule applied','No estimates. A figure appears only if a college document states it, or if the same figure is '
    'confirmed in the 2026-27 admission documents of two or more colleges in that state.'),
   ('Maharashtra','Rs 1,67,300/yr tuition for 2026-27, stated independently in the admission brochures of GMC Bhandara, '
    'GMC Ratnagiri, GMC Baramati, GMC Nandurbar, VNGMC Yavatmal and VMGMC Miraj. Series in the same documents: '
    'Rs 78,000 (2018-19) -> Rs 94,400 (2020-21) -> Rs 1,03,900 (2021-22) -> Rs 1,38,300 (2024-25) -> '
    'Rs 1,52,100 (2025-26) -> Rs 1,67,300 (2026-27), per MGIMS academic calendars and college brochures. '
    'Rs 83,650 recurs as the 50% concession rate. The doubling in five years is why third-party figures are all stale.'),
   ('AIQ candidates pay full','GMC Ratnagiri fee table: "All India Candidates (both Open & Reserve category) - 2 DDs, '
    'Rs 1,67,300 + Rs 14,000". Maharashtra reserved categories and low-income Maharashtra women have tuition met by state '
    'scholarship; a General AIQ candidate does not.'),
   ('Latur and Akola','Cannot be confirmed from their own notifications because they publish no documents at all - which is '
    'exactly why they were the two cleanest privacy results. Their rows carry the state figure, marked as not confirmed at '
    'the college.'),
   ('Rejected sources','Aggregator sites returned Rs 1,62,100 / Rs 70,500 / Rs 6,92,000 for the same Maharashtra colleges, '
    'mutually inconsistent and untraceable to a notification. A West Bengal search returned Rs 2-4 lakh per semester for '
    'government colleges, which is private/NRI tuition misattributed. Neither was used. DMER Maharashtra publishes no MBBS '
    'fee link and CET Cell publishes only engineering fee notices.'),
   ('Tamil Nadu','GMC Karur publishes a dedicated AIQ_FEES.pdf stating Rs 6,000 tuition; GMC Namakkal FEES-structure.pdf '
    'states Rs 3,000/Rs 6,000. These apply to AIQ candidates directly.'),
   ('Other confirmed','NEIGRIHMS Shillong Rs 3,000 (MBBS Prospectus 2026). GMC Anantnag Rs 75,000. GMC Chhindwara Rs 1,00,000.'),
   ('Calendar column','"Yes" means an academic calendar, almanac or master time-table is published. Checked by fetching each '
    'site plus 15 calendar-specific paths and following calendar-labelled links and documents.'),
   ('Blank cells','57 of the 135 colleges have no site reachable from this environment (geo-blocked state portals), so they '
    'have no fee document and no calendar to check. Those rows say so rather than showing a blank that could read as "nothing published".')]
for j,(h,w) in enumerate([('Topic',26),('Detail',150)],start=1):
    c=ns.cell(row=1,column=j,value=h); c.font=HDR; c.fill=HFIL
    ns.column_dimensions[get_column_letter(j)].width=w
for i,(a,b) in enumerate(N,start=2):
    ns.cell(row=i,column=1,value=a).alignment=WRAP
    ns.cell(row=i,column=2,value=b).alignment=WRAP

wb.save(os.path.join(HERE,'AIQ_all_colleges_fees_calendar.xlsx'))
print('rows:',len(out))
print('calendar Yes:',sum(1 for o in out if o['cal']=='Yes'),
      '| No:',sum(1 for o in out if o['cal']=='No'),
      '| unreachable/unchecked:',sum(1 for o in out if o['cal'] not in ('Yes','No')))
print('fee from college doc:',sum(1 for o in out if o['basis'].startswith('College document')),
      '| state-confirmed:',sum(1 for o in out if 'NOT confirmed at this college' in o['basis']),
      '| NOT CONFIRMED:',sum(1 for o in out if o['basis'].startswith('NOT CONFIRMED') or o['basis'].startswith('Site not reachable')))
print('under 40k:',sum(1 for o in out if o['under']=='Yes'))
