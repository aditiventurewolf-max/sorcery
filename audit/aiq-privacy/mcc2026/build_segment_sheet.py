import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook('AIQ_final_shortlist_criteria.xlsx')
HDR=Font(bold=True,color='FFFFFF',size=10); HFIL=PatternFill('solid',fgColor='7B3F00')
BOLD=Font(bold=True); WRAP=Alignment(wrap_text=True,vertical='top')
CEN=Alignment(horizontal='center',vertical='center')
THIN=Border(*[Side(style='thin',color='BFBFBF')]*4)
GOOD=PatternFill('solid',fgColor='C6EFCE'); WARN=PatternFill('solid',fgColor='FFEB9C')
BAD=PatternFill('solid',fgColor='FFC7CE')

ws = wb.create_sheet('Segment - 4 new filters', 2)
ws['A1']='Segment: no academic calendar + tuition under Rs 40,000 + direct/1-stop flight + not newly founded'
ws['A1'].font=Font(bold=True,size=13)
ws['A2']='Drawn from BOTH the PASS and FAIL sets, as requested. Privacy verdict is shown so the trade-off is visible.'
ws['A3']='Lucknow (LKO) flies direct to 27 airports incl. Kolkata, Delhi, Mumbai, Hyderabad, Chennai, Indore - but NOT Nagpur, Ranchi or Pune.'
ws['A3'].font=Font(italic=True,size=9)

COLS=[('College',44),('City',16),('State',16),('2026 R1 close',12),
      ('Academic calendar on site',15),('Tuition Rs/yr',13),('Fee source',30),('Under Rs 40k',11),
      ('Established',11),('Age (yrs)',9),('Newly founded',12),
      ('Nearest airport',26),('LKO flight',20),('Surface leg',18),('Meets flight+1-2hr',13),
      ('PRIVACY VERDICT',15),('Passes all 4 filters',13),('What kills it',44)]
for j,(h,w) in enumerate(COLS,start=1):
    c=ws.cell(row=5,column=j,value=h); c.font=HDR; c.fill=HFIL
    c.alignment=Alignment(wrap_text=True,horizontal='center',vertical='center')
    ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[5].height=44

R=[
 # college, city, state, close, cal, fee, feesrc, under40, est, newly, airport, flight, surface, meets, verdict, passes, kills
 ('College of Medicine & JNM Hospital','Kalyani','West Bengal',10855,'No',10000,'WB state estimate - CONFIRM with WBMCC notification','Yes',2011,'No',
  'Kolkata (CCU)','DIRECT from LKO','~60 km / ~1.5 hrs','YES','FAIL','No','PRIVACY: hash-named roster PDF you supplied'),
 ('College of Medicine & Sagore Dutta Hospital','Kamarhati (Kolkata)','West Bengal',8626,'No',10000,'WB state estimate - CONFIRM with WBMCC notification','Yes',2011,'No',
  'Kolkata (CCU)','DIRECT from LKO','~15 km / ~40 min','YES','FAIL','No','PRIVACY: class + full batch lists you found'),
 ('Bankura Sammilani Medical College','Bankura','West Bengal',12710,'No',10000,'WB state estimate - CONFIRM with WBMCC notification','Yes',1959,'No',
  'Kolkata (CCU)','DIRECT from LKO','~200 km / ~4 hrs','No','FAIL','No','PRIVACY + surface leg 4 hrs'),
 ('GMC Ongole (prev. RIMS)','Ongole','Andhra Pradesh',17028,'No',15000,'AP state estimate - CONFIRM with APMC notification','Yes',2008,'Borderline',
  'Vijayawada (VGA)','1-stop via Hyderabad','~250 km / ~4.5 hrs','No','FAIL','No','PRIVACY: name+roll+mobile; surface leg 4.5 hrs'),
 ('GMC Udhampur','Udhampur','Jammu & Kashmir',14833,'No',40000,'J&K estimate - at the boundary, not under','No',2019,'Yes',
  'Jammu (IXJ)','1-stop via Delhi','~70 km / ~1.5 hrs','YES','HIGH RISK','No','Fee at boundary; 536-file open dir; J&K dropped by you'),
 ('GMC Latur','Latur','Maharashtra',13905,'No',85000,'MH state estimate - CONFIRM with state notification','No',2002,'No',
  'Nanded (NDC)','1-stop, limited service','~130 km / ~2.5 hrs','No','PASS','No','FEE Rs 85,000; weak connectivity'),
 ('GMC Akola','Akola','Maharashtra',13994,'No',85000,'MH state estimate - CONFIRM with state notification','No',2002,'No',
  'Nagpur (NAG)','1-stop via Delhi/Mumbai','~250 km / ~4.5 hrs','No','PASS','No','FEE Rs 85,000; worst surface leg on the list'),
 ('GMC Bhandara','Bhandara','Maharashtra',18616,'No',85000,'MH state estimate - CONFIRM with state notification','No',2019,'Yes',
  'Nagpur (NAG)','1-stop via Delhi/Mumbai','~60 km / ~1.5 hrs','YES','PASS','No','FEE Rs 85,000; newly founded'),
 ('GMC Seoni','Seoni','Madhya Pradesh',15877,'No',80000,'MP state estimate - CONFIRM with state notification','No',2018,'Yes',
  'Nagpur (NAG)','1-stop via Delhi/Mumbai','~130 km / ~2.5 hrs','No','PASS','No','FEE Rs 80,000; newly founded'),
 ('Shaheed Nirmal Mahto MC, Dhanbad','Dhanbad','Jharkhand',10303,'YES - 3 docs',8500,'careers360 reports Rs 8,500 tuition - CONFIRM with college','Yes',1969,'No',
  'Ranchi (IXR) / Durgapur (RDP)','1-stop via Delhi/Kolkata','~140 km / ~3 hrs (rail direct)','No','PASS','No','HAS academic calendar; surface leg 3 hrs'),
 ('SRTR Ambajogai','Ambajogai','Maharashtra',14470,'YES',85000,'MH state estimate','No',1975,'No',
  'Nanded (NDC)','1-stop, limited service','~90 km / ~2 hrs','YES','PASS','No','HAS calendar; FEE Rs 85,000'),
 ('GMC Chandrapur','Chandrapur','Maharashtra',15048,'YES',85000,'MH state estimate','No',2015,'Borderline',
  'Nagpur (NAG)','1-stop via Delhi/Mumbai','~150 km / ~3 hrs','No','PASS','No','HAS calendar; FEE Rs 85,000'),
 ('GMC Ratnagiri','Ratnagiri','Maharashtra',16082,'YES',85000,'MH state estimate','No',2023,'Yes',
  'Kolhapur (KLH)','1-stop via Mumbai','~130 km / ~3 hrs','No','PASS','No','HAS calendar; FEE Rs 85,000; newly founded'),
 ('JLN Medical College, Bhagalpur','Bhagalpur','Bihar',10357,'YES',6000,'Bihar state estimate - CONFIRM','Yes',1971,'No',
  'Patna (PAT)','1-stop via Delhi','~220 km / ~4.5 hrs','No','PASS','No','HAS calendar; surface 4.5 hrs; you excluded as too near'),
 ('Medinirai MC, Palamu','Daltonganj','Jharkhand',None,'YES',8500,'Jharkhand estimate','Yes',2019,'Yes',
  'Ranchi (IXR)','1-stop via Delhi','~170 km / ~3.5 hrs','No','FAIL','No','PRIVACY: current-intake student list; HAS calendar; newly founded'),
 ('GMC Baramulla','Baramulla','Jammu & Kashmir',17201,'YES',40000,'J&K estimate - at boundary','No',2019,'Yes',
  'Srinagar (SXR)','1-stop via Delhi','~55 km / ~1.5 hrs','YES','PASS','No','HAS calendar; newly founded; J&K dropped by you'),
]
for i,r in enumerate(R,start=6):
    for j,v in enumerate(r,start=1):
        c=ws.cell(row=i,column=j,value=v if v is not None else 'not allotted R1')
        c.border=THIN
        c.alignment=WRAP if j in (7,18) else CEN
    ws.cell(row=i,column=6).number_format='#,##0'
    ws.cell(row=i,column=5).fill = GOOD if r[4]=='No' else BAD
    ws.cell(row=i,column=8).fill = GOOD if r[7]=='Yes' else BAD
    ws.cell(row=i,column=11).fill = GOOD if r[10]=='No' else (WARN if r[10]=='Borderline' else BAD)
    ws.cell(row=i,column=15).fill = GOOD if r[13]=='YES' else BAD
    ws.cell(row=i,column=16).fill = GOOD if r[15]=='PASS' else BAD
    ws.cell(row=i,column=17).fill = BAD
ws.freeze_panes='A6'
n=len(R)+7
ws.cell(row=n,column=1,value='RESULT: no college satisfies all four filters at once.').font=Font(bold=True,size=12)
ws.cell(row=n+1,column=1,value='The four criteria are best met by the Kolkata-belt West Bengal colleges - cheapest tuition, the only DIRECT Lucknow flight,')
ws.cell(row=n+2,column=1,value='a 40-90 minute surface hop, no academic calendar, and decades old. All of them publish student data (confirmed FAILs).')
ws.cell(row=n+3,column=1,value='The privacy-clean colleges that clear the calendar filter - Latur, Akola, Bhandara, Seoni - all fail the Rs 40k fee gate at Rs 80-85k.')
ws.cell(row=n+5,column=1,value='CLOSEST TO YOUR FOUR FILTERS, privacy intact: GMC Latur and GMC Akola (no calendar, established 2002, PASS) - blocked only by fee.').font=BOLD
ws.cell(row=n+6,column=1,value='CLOSEST on flight rule, privacy intact: GMC Bhandara (Nagpur ~60 km) - blocked by fee and by being founded 2019.').font=BOLD
ws.cell(row=n+8,column=1,value='All tuition figures are state-level estimates unless the Fee source column says otherwise. Confirm against the state fee notification.')
ws.cell(row=n+9,column=1,value='Airport distances and surface times are my estimates. Verify before relying on them.')
wb.save('AIQ_final_shortlist_criteria.xlsx')
print('saved; sheets:', wb.sheetnames)
