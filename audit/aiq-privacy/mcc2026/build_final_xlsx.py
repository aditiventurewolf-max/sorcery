import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
HDR  = Font(bold=True, color='FFFFFF', size=10)
HFIL = PatternFill('solid', fgColor='1F4E79')
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')
CEN  = Alignment(horizontal='center', vertical='center')
THIN = Border(*[Side(style='thin', color='BFBFBF')]*4)
GOOD = PatternFill('solid', fgColor='C6EFCE')
WARN = PatternFill('solid', fgColor='FFEB9C')
BAD  = PatternFill('solid', fgColor='FFC7CE')
GREY = PatternFill('solid', fgColor='F2F2F2')

# ---------------------------------------------------------------- Weights
ws = wb.active; ws.title = 'Weights & Method'
ws['A1'] = 'Weights — edit column B, the shortlist score recalculates'
ws['A1'].font = Font(bold=True, size=13)
ws['A3'] = 'Scoring: every score column is 1-5 where 5 = best for you.'
ws['A4'] = 'Staleness 5 = site is most neglected. Distance 5 = farthest from Lucknow. Fee 5 = cheapest.'
ws['A5'] = 'Connectivity 5 = airport within a 1-2 hr train/bus hop, per your "flight + 1-2 hrs" rule.'
rows = [('Info exposure (nothing student-related online)', 0.40, 'L — Info-exposure score'),
        ('Website staleness (un-updated is good)',          0.15, 'O — Staleness score'),
        ('Connectivity (flight + 1-2 hr hop)',              0.15, 'Y — Connectivity score'),
        ('Distance from Lucknow',                           0.10, 'J — Distance score'),
        ('Fees',                                            0.10, 'U — Fee score'),
        ('Hindi spoken',                                    0.10, 'P — Hindi score')]
ws['A7'] = 'Factor'; ws['B7'] = 'Weight'; ws['C7'] = 'Score column'
for c in 'ABC': ws[f'{c}7'].font = HDR; ws[f'{c}7'].fill = HFIL
for i,(n,w,col) in enumerate(rows, start=8):
    ws[f'A{i}'] = n; ws[f'B{i}'] = w; ws[f'C{i}'] = col
    ws[f'B{i}'].number_format = '0%'
ws['A14'] = 'TOTAL'; ws['A14'].font = BOLD
ws['B14'] = '=SUM(B8:B13)'; ws['B14'].font = BOLD; ws['B14'].number_format = '0%'
ws['A16'] = 'Alternative: connectivity-first weighting (copy these into B8:B13 to switch)'
ws['A16'].font = BOLD
alt = [('Info exposure', 0.30), ('Website staleness', 0.10), ('Connectivity', 0.30),
       ('Distance from Lucknow', 0.10), ('Fees', 0.10), ('Hindi spoken', 0.10)]
for i,(n,w) in enumerate(alt, start=17):
    ws[f'A{i}'] = '   ' + n; ws[f'B{i}'] = w; ws[f'B{i}'].number_format = '0%'
ws['A23'] = 'Under connectivity-first: Dhanbad stays #1 (4.00), Bhandara rises to #2 (3.80), Akola drops from #2 to #5.'
ws['A25'] = 'NOTE: km and travel times are my estimates from public sources — verify before booking.'
ws['A26'] = 'NOTE: bond and PG-quota rules are revised annually. Confirm in the current counselling brochure.'
ws['A27'] = 'NOTE: fees are state tuition estimates per year, excluding hostel/mess/caution deposits.'
ws.column_dimensions['A'].width = 52; ws.column_dimensions['B'].width = 12; ws.column_dimensions['C'].width = 26

# ---------------------------------------------------------------- Shortlist
# college, city, state, close26, open26, seats26, close25, km, dscore,
# docs_opened, iexp, whatpublished, newest, stale, hindi, hindinote, fee, feeband, feedoc, fescore,
# airport, airport_km, airport_hrs, connscore, bond, pg
D = [
 ('Shaheed Nirmal Mahto Medical College & Hospital','Dhanbad','Jharkhand',10303,8204,5,9623,
  900,3, 31,4,'No roster. 31 docs opened, no open directories.', '2026',3,
  5,'Hindi is the working language', 40000,'Low','N',5,
  'Durgapur (RDP) / Ranchi (IXR)',90,'~2 hrs',4,
  'Bond applies','MBBS-in-state generally qualifies'),
 ('Government Medical College, Latur','Latur','Maharashtra',13905,9885,7,12233,
  1300,5, 0,5,'Zero documents of any kind on the whole site.','2026',3,
  3,'Marathi; Hindi widely understood',85000,'High','N',1,
  'Nanded (NDC) / Latur (LTU, limited)',130,'~2.5 hrs',2,
  '~1 yr bond','MBBS-in-state generally qualifies'),
 ('Government Medical College, Akola','Akola','Maharashtra',13994,11441,8,12815,
  1000,4, 1,5,'Docs behind api.gmcakola.in — store empty on every parameter.','2026',5,
  3,'Marathi; Hindi widely understood',85000,'High','N',1,
  'Nagpur (NAG)',250,'~4.5 hrs',1,
  '~1 yr bond','MBBS-in-state generally qualifies'),
 ('Swami Ramanand Teerth Rural Medical College, Ambajogai','Ambajogai','Maharashtra',14470,12368,6,12681,
  1250,4, 388,3,'388 docs opened. MUHS disclosure filings only (faculty/beds/dean).','2026',3,
  3,'Marathi; Hindi understood',85000,'High','Y',1,
  'Nanded (NDC) / Latur (LTU, limited)',90,'~2 hrs',2,
  '~1 yr bond','MBBS-in-state generally qualifies'),
 ('Government Medical College, Chandrapur','Chandrapur','Maharashtra',15048,8176,7,12027,
  1050,4, 186,3,'186 docs opened. Flags were a Marathi circular + library journal list.','2026',3,
  3,'Marathi/Hindi belt; Hindi common',85000,'High','N',1,
  'Nagpur (NAG)',150,'~3 hrs',2,
  '~1 yr bond','MBBS-in-state generally qualifies'),
 ('Government Medical College, Seoni','Seoni','Madhya Pradesh',15877,11893,7,12759,
  800,3, 107,4,'107 docs opened. Only flag was faculty/resident AEBAS attendance.','2026',3,
  5,'Hindi is the state language',80000,'High','N',2,
  'Nagpur (NAG) / Jabalpur (JLR)',130,'~2.5 hrs',3,
  '~1 yr bond','MBBS-in-state generally qualifies'),
 ('Government Medical College, Ratnagiri','Ratnagiri','Maharashtra',16082,15646,2,12093,
  1500,5, 75,4,'75 docs opened. Flags were admission brochures with blank fields.','2026',3,
  3,'Marathi; Hindi less common in Konkan',85000,'High','Y',1,
  'Kolhapur (KLH) / Ratnagiri (RTC, limited)',130,'~3 hrs',2,
  '~1 yr bond','MBBS-in-state generally qualifies'),
 ('Government Medical College, Bhandara','Bhandara','Maharashtra',18616,16568,5,12648,
  1000,4, 25,4,'116-path battery + 25 docs. Flags were admission instruction brochures.','2026',3,
  3,'Marathi/Hindi belt; Hindi common',85000,'High','Y',1,
  'Nagpur (NAG)',60,'~1.5 hrs',5,
  '~1 yr bond','MBBS-in-state generally qualifies'),
]

ws = wb.create_sheet('Final shortlist')
COLS = [('Rank by score',14),('College',46),('City',14),('State',16),
        ('2026 R1 closing AIR',13),('2026 opening AIR',12),('R1 Gen seats',9),('2025 R1 close',12),
        ('Km from Lucknow',12),('Distance score',10),
        ('Docs opened',10),('Info-exposure score',12),('What the site publishes',52),
        ('Newest year on site',11),('Staleness score',10),
        ('Hindi spoken score',10),('Hindi note',34),
        ('Tuition Rs/yr',12),('Fee band',9),('Fee doc on site',9),('Fee score',9),
        ('Nearest airport',34),('Airport km',9),('Airport surface time',13),('Connectivity score',11),
        ('UG service bond',16),('PG state quota for AIQ entrant',34),
        ('WEIGHTED SCORE',13)]
for j,(h,w) in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=j, value=h)
    c.font = HDR; c.fill = HFIL; c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[1].height = 46

for i,r in enumerate(D, start=2):
    (coll,city,st,c26,o26,s26,c25,km,ds,docs,iexp,what,newest,stale,
     hin,hnote,fee,fband,fdoc,fsc,ap,apkm,aphrs,conn,bond,pg) = r
    vals = [None,coll,city,st,c26,o26,s26,c25,km,ds,docs,iexp,what,newest,stale,
            hin,hnote,fee,fband,fdoc,fsc,ap,apkm,aphrs,conn,bond,pg,None]
    for j,v in enumerate(vals, start=1):
        cell = ws.cell(row=i, column=j, value=v)
        cell.border = THIN
        if j in (13,17,22,26,27): cell.alignment = WRAP
        else: cell.alignment = CEN
    ws.cell(row=i, column=18).number_format = '#,##0'
    for j,sc in ((10,ds),(12,iexp),(15,stale),(16,hin),(21,fsc),(25,conn)):
        ws.cell(row=i, column=j).fill = GOOD if sc>=4 else (WARN if sc==3 else BAD)
    ws.cell(row=i, column=28).value = (
        f"=L{i}*'Weights & Method'!$B$8+O{i}*'Weights & Method'!$B$9"
        f"+Y{i}*'Weights & Method'!$B$10+J{i}*'Weights & Method'!$B$11"
        f"+U{i}*'Weights & Method'!$B$12+P{i}*'Weights & Method'!$B$13")
    ws.cell(row=i, column=28).number_format = '0.00'
    ws.cell(row=i, column=28).font = BOLD
    ws.cell(row=i, column=1).value = f'=RANK(AB{i},$AB$2:$AB${len(D)+1})'
    ws.cell(row=i, column=1).font = BOLD
    if city == 'Dhanbad':
        for j in (2,3,4,5): ws.cell(row=i,column=j).font = BOLD
ws.freeze_panes = 'B2'
n = len(D)+1
ws.cell(row=n+2, column=2, value='Only Dhanbad (10,303) closes inside your 8,000-13,000 band. The rest close 900-5,600 ranks beyond it,').font = BOLD
ws.cell(row=n+3, column=2, value='which means they are reachable with headroom at that rank, not out of reach.')
ws.cell(row=n+5, column=2, value='Jammu & Kashmir removed per your instruction (GMC Baramulla, closing 17,201, was otherwise clean).')

# ---------------------------------------------------------------- Evidence
ws = wb.create_sheet('Verification evidence')
EV = [('Dhanbad','www.snmmcdhn.org','88 URLs crawled, 31 docs opened, recursive open-dir walk found NO open directories, 35 fixed disclosure paths probed','Clean. Least-checked of the list — worth your own spot-check.'),
      ('Latur','gmclatur.org','17 URLs = full coverage of an 11-page brochure site. Zero documents. Re-walk: no open dirs.','Clean. Verified the 17-URL crawl is complete, not shallow.'),
      ('Akola','gmcakola.in','Docs served from api.gmcakola.in/api/Auth/get-all-WithoutAuth. Enumerated Type 0-12 x IsWeb true/false: empty on all.','Clean, confirmed at the source rather than by absent evidence.'),
      ('Ambajogai','srtrmca.org','56 URLs, 388 docs opened. 3 flags = MUHS mandatory-disclosure filings; 2025-26 one is an unfilled template.','Clean. Large doc surface but no rosters.'),
      ('Chandrapur','gmcchandrapur.org','83 URLs, 186 docs opened. 2 flags = Marathi circular + library journal subscription list.','Clean.'),
      ('Seoni','gmcseoni.edu.in','95 URLs, 107 docs opened. 1 flag = faculty/resident AEBAS attendance (professor x51, resident x28).','Clean.'),
      ('Ratnagiri','gmcratnagiri.in','110 URLs, 75 docs opened. 3 flags = admission guideline brochures with blank signature/father fields.','Clean.'),
      ('Bhandara','gmcbhandara.edu.in','116 fixed paths + 110-page best-of-N crawl earlier; 25 docs opened now. 2 flags = admission instruction brochures.','Clean. Most deeply checked on the list.')]
for j,(h,w) in enumerate([('College',16),('Host',26),('What was checked',96),('Verdict',60)], start=1):
    c=ws.cell(row=1,column=j,value=h); c.font=HDR; c.fill=HFIL
    ws.column_dimensions[get_column_letter(j)].width=w
for i,r in enumerate(EV, start=2):
    for j,v in enumerate(r, start=1):
        c=ws.cell(row=i,column=j,value=v); c.alignment=WRAP; c.border=THIN

# ---------------------------------------------------------------- Failed
ws = wb.create_sheet('FAILED - avoid')
F = [(11791,'GMC Shivpuri','Madhya Pradesh','files/BATCH 2023-24.pdf, BATCH 2024-25.pdf','Name of Student + NEET percentile, serial rows'),
     (13813,'Lt. L.A.M. GMC Raigarh','Chhattisgarh','storage/grievances.csv, storage/feedback.csv','4 records marked Confidential=Yes with name, email, phone, complaint text'),
     (15154,'VNGMC Yavatmal','Maharashtra','studentsection/U2015,U2016,U2019,U2020.pdf + 7 exam results','Name of the Student / M-F / Date of admission / reservation Category'),
     (17028,'GMC Ongole','Andhra Pradesh','UG-STUDENTS-LIST-2022-BATCH.pdf + 12 more','Name + roll number + 10-digit mobile; monthly PG stipend lists'),
     (7384,'GMC Baramati','Maharashtra','brcp/studet_list_mbbs/ (3 files)','Batch 2020-21, 2021-22, 2022-23. Also closes below your band.'),
     (None,'Medinirai MC, Palamu','Jharkhand','Student-List-25-26.pdf, MBBS-BATCH-2023.pdf, Student-List-2019.pdf','CURRENT intake published. 89 and 86 row lists.'),
     (12178,'GMC Jalgaon','Maharashtra','assets/pdf/admission/ug/UG Data-2025-26.pdf','Current 2025-26 intake, 2.2 MB scan'),
     (9749,'GMC Nashik','Maharashtra','NMC Information page','Roster page behind non-matching link text'),
     (10855,'COM & JNM Hospital, Kalyani','West Bengal','assets/pdf/<hash>.pdf','Hash-named roster PDF'),
     (8626,'COM & Sagore Dutta Hospital','West Bengal','class + full batch lists','Slow server hid them at short timeouts'),
     (9122,'Kakatiya MC, Warangal','Telangana','CSV data files, 5 sessions incl. 2025-26','Name paired with NEET RANK — most severe find of the audit'),
     (10873,'GMC The Nilgiris','Tamil Nadu','batch-2025-26-list-of-students.pdf','Current intake'),
     (12751,'GMC Datia','Madhya Pradesh','B.1.11 disclosure page','batch2025, 136 rows'),
     (10313,'GMC Jhunjhunu','Rajasthan','Scribd re-host','100 students, names + fathers names'),
     (8812,'NAMO Silvassa','DNH & DD','5 batch PDFs','Self-hosted'),
     (9265,'S V Medical College, Tirupati','Andhra Pradesh','B.1.11 page','240 rows against 240 sanctioned seats'),
     (9640,'Hassan Inst. of Medical Sciences','Karnataka','Scribd re-host','2014 batch')]
for j,(h,w) in enumerate([('2026 R1 close',13),('College',40),('State',18),('Where',52),('What is exposed',62)], start=1):
    c=ws.cell(row=1,column=j,value=h); c.font=HDR; c.fill=HFIL
    ws.column_dimensions[get_column_letter(j)].width=w
for i,r in enumerate(F, start=2):
    for j,v in enumerate(r, start=1):
        c=ws.cell(row=i,column=j,value=v if v is not None else 'not allotted in R1')
        c.alignment=WRAP; c.border=THIN
        if j==2: c.fill=BAD

wb.save('AIQ_final_shortlist_criteria.xlsx')
print('saved')
