import csv, re, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
M='/home/user/sorcery/audit/aiq-privacy/mcc2026/all_2026_aiq.csv'
rows=[(int(r['closing_AIR_2026_R1']),int(r['opening_AIR']),int(r['R1_gen_seats']),r['institute'])
      for r in csv.DictReader(open(M))]

# ---- direct non-stop from Lucknow (confirmed Aug 2026), with airport + surface leg
DIRECT={
 'delhi':('Delhi (DEL)','~15 km / ~40 min','Most frequent route from LKO; many daily options'),
 'noida':('Noida Intl (DXN)','~60 km','Newer airport, thinner schedule'),
 'kolkata':('Kolkata (CCU)','varies','Multiple daily IndiGo'),
 'hyderabad':('Hyderabad (HYD)','~35 km','Multiple daily'),
 'secunderabad':('Hyderabad (HYD)','~35 km','Multiple daily'),
 'indore':('Indore (IDR)','~10 km / ~25 min','Daily'),
 'raipur':('Raipur (RPR)','~15 km / ~30 min','IndiGo non-stop'),
 'patna':('Patna (PAT)','~10 km / ~30 min','IndiGo non-stop'),
 'pune':('Pune (PNQ)','varies','IndiGo / Air India Express'),
 'bengaluru':('Bengaluru (BLR)','varies','IndiGo / Akasa / AIX'),
 'bangalore':('Bengaluru (BLR)','varies','IndiGo / Akasa / AIX'),
 'mumbai':('Mumbai (BOM)','varies','Multiple daily'),
 'goa':('Goa (GOX/GOI)','~30-35 km','Direct, seasonal frequency'),
 'panaji':('Goa (GOX/GOI)','~30-35 km','Direct, seasonal frequency'),
 'bambolim':('Goa (GOX/GOI)','~30 km','Direct, seasonal frequency'),
 'guwahati':('Guwahati (GAU)','varies','Direct'),
 'jaipur':('Jaipur (JAI)','~12 km','Direct'),
 'dehradun':('Dehradun (DED)','~25 km','Direct'),
 'chandigarh':('Chandigarh (IXC)','~12 km','Direct'),
 'chennai':('Chennai (MAA)','varies','Direct'),
 'bhopal':('Bhopal (BHO)','~15 km','Direct'),
 'jharsuguda':('Jharsuguda (JRG)','varies','Star Air direct'),
 'kishangarh':('Kishangarh (KQH)','~30 km to Ajmer','Star Air direct'),
 'ajmer':('Kishangarh (KQH)','~30 km','Star Air direct'),
}
# ---- privacy findings from this audit (token -> (verdict, newest, detail))
PRIV={
'raipur':('FAIL - stale','2017','813 docs opened, 11 rosters: First_MBBS.pdf (52 rows), MBBS_II_dec2014.pdf (45), UG2017.pdf and U2017.pdf (27 each), plus PG result lists. Newest UG list 2017, nothing after.'),
'goa medical':('FAIL - current','2022-23','5 "Final Roll No. of MBBS" files 2018-19 to 2022-23, each a Dean letter listing admitted 1st-MBBS students by NAME and GENDER (61-99 rows); 4 year-files; paramedical lists. /nmc is a 480 KB disclosure hub. 10 hits / 64 docs.'),
'baramati':('FAIL - current','2022-23','Directory literally named brcp/studet_list_mbbs/ holding batch 2020-21, 2021-22, 2022-23, each with a student-name header.'),
'jabalpur':('FAIL - deeply hidden, stale','2017','Angular SPA. No student/document/downloads/academic/calendar endpoint (all 404). /api/notice and /api/results return count 0. News store newest 2020-01-21. Single roster final17.pdf (138 rows, "Name of Student") reachable ONLY by POST to api.nscbmc.ac.in/api/news.'),
'mandya':('Clean on this pass','n/a','130 URLs crawled, 0 documents of any kind, no open directories. NOT confirmed against the SPA trap - treat as unverified rather than clean.'),
'rg kar':('FAIL - stale','Jan 2012','/pdf/student-list-jan2012.pdf (53 rows), student-list-jan2010.pdf, student-list-categorywise-2011.pdf. Whole site frozen 2010-2012. Current official site rgkarmch.in is geo-blocked, so this covers only the reachable legacy site.'),
'sagore dutta':('FAIL','unknown','Class lists and full batch lists; only visible at long timeouts (slow server).'),
'sagar dutta':('FAIL','unknown','Class lists and full batch lists; only visible at long timeouts.'),
'ajmer':('FAIL - current','current','Rajasthan state portal template ships a POPULATED "Information under clause B.1.11 - List of Students and Sanctioned Intake Capacity" page. Confirmed for JLN Ajmer and SMS Jaipur.'),
'rewa':('Clean but largest surface','n/a','110 URLs, 1,611 documents - by far the biggest document surface of any cleared college. Weakest of the clean results.'),
'kakatiya':('FAIL - most severe','2025-26','CSV data files pairing NAME with NEET RANK across 5 sessions including the current 2025-26; PG lists add NEET roll number.'),
'tirupati':('FAIL - current','2025-26','1,776 docs, 7 roster hits including a 2025-26 PG prospectus; B.1.11 page links UG name lists, 240 rows against 240 sanctioned seats.'),
'bilaspur':('CLEAN - best result','n/a','755 documents opened, ZERO roster hits. More docs opened than any other cleared college (Ambajogai 388, Chandrapur 186). Publishes an academic calendar.'),
'nashik':('FAIL - current','2025-26','MBBS Student List Batch 2025.pdf (Final Retention Data 2025-26) and Batch 2024.pdf from its NMC Information page.'),
'burdwan':('FAIL - in your face, but stale','2016-17','/admitted-candidates.php is a LIVE HTML TABLE, ~150 rows, columns: S.No | Roll No | NEET AIR | Candidate Name | Subject | Allotted Category | Candidate Category | Reporting Date. Plus Admitted_Students_of_2015.pdf, UG_MBBS_2014List_for_MCI.pdf, MCI_Data_2016-17.pdf (117 rows), 2015-16.pdf. Most identifying exposure in the audit - name + NEET AIR + roll no together.'),
'dhanbad':('CLEAN - best verified','n/a','FULL BATTERY: 413 fixed disclosure paths, 89-URL crawl, recursive open-directory re-walk found NO open dirs, all 31 docs opened, 2 scanned PDFs OCR-read (subject/exam schedules, no names). Zero hits.'),
'kalyani':('FAIL','unknown','Hash-named roster PDF at /assets/pdf/<hash>.pdf - no meaningful link text.'),
'nilgiris':('FAIL - current','2025-26','/page/list-of-student serves per-batch name PDFs including batch-2025-26-list-of-students.pdf (header NAME, serials to 150).'),
'karur':('Unverified - merit list page','unknown','Second domain tiruppurmedicalcollege-style: gmchtiruppur/karurgmc pattern. Fee doc found (AIQ_FEES.pdf). Workbook recorded no documents, but a student-merit-list URL exists on the sister domain - treat as unverified.'),
'namakkal':('Unverified - merit list page','unknown','namakkalmedicalcollege.in/page/student-merit-list recorded in evidence column while the summary said no documents. Contradiction in my own data - unverified.'),
'shivpuri':('FAIL - current','2024-25','files/BATCH 2023-24.pdf and BATCH 2024-25.pdf - "Name of Student" + NEET percentile per row; PG merit list with names and percentiles.'),
'jalgaon':('FAIL - current','2025-26','assets/pdf/admission/ug/UG Data-2025-26.pdf - current intake, 2.2 MB scan, in a directory linked from nowhere crawlable.'),
'krishnagiri':('PASS VOID','unknown','My earlier clearance was verified against the WRONG SITE (an Indonesian midwives association). Never actually checked.'),
'neigrihms':('FAIL','unknown','Confirmed FAIL in the audit. No academic calendar. Fee Rs 3,000 confirmed from its MBBS Prospectus 2026.'),
'shillong':('FAIL','unknown','Confirmed FAIL. Fee Rs 3,000 confirmed from MBBS Prospectus 2026.'),
'employees state insurance':('UNVERIFIED - structurally low','n/a','ESIC colleges run on ONE locked central template (*.esic.gov.in) with no student-list module and no per-college admin who could add one. All ESIC sites are geo-blocked from this environment so I could not confirm. User checking from India found no student list, consistent with the template explanation. NMC MSMER portal proforma still needs checking.'),
'esic':('UNVERIFIED - structurally low','n/a','ESIC central template has no student-list module. Geo-blocked here; user found none from India.'),
'nalanda':('FAIL - current','2024-25','nmchpatna.ac.in publishes 18 sessions of student lists, 2006-07 through 2024-25. This was the audit brief exemplar used to validate the detector.'),
'jhunjhunu':('FAIL','2024-25','Scribd re-host: MBBS batch 2024-25, 100 students, names + fathers names.'),
'silvassa':('FAIL','current','5 batch PDFs self-hosted (brief exemplar, re-verified).'),
'hassan':('FAIL - stale','2014','Scribd re-host of the 2014 batch.'),
'datia':('FAIL - current','2025','B.1.11 disclosure page, batch2025, 136 rows.'),
'miraj':('Unresolved','n/a','Scanned-image UG admission PDF with no extractable text - PROVISIONAL pending OCR, never cleared.'),
}
# ---- fees: only document-backed figures
FEE={'karur':(6000,'College doc: karurgmc.ac.in AIQ_FEES.pdf - AIQ-specific'),
 'namakkal':(6000,'College doc: namakkalmedicalcollege.in FEES-structure.pdf'),
 'neigrihms':(3000,'College doc: MBBS Prospectus 2026'),
 'shillong':(3000,'College doc: MBBS Prospectus 2026'),
 'bilaspur':(40000,'College doc - OVER the Rs 35,000 ceiling'),
 'baramati':(167300,'Maharashtra 2026-27, own brochure'),
 'miraj':(167300,'Maharashtra state figure (8 colleges concur)'),
 'nashik':(152100,'Own brochure (2025-26 figure)'),
 'jalgaon':(167300,'Maharashtra state figure'),
}

# ---- proper display names + city extraction + nearest airport
_BAD=re.compile(r'college|hospital|institute|road|marg|nagar|dist|\bpo\b|pin|near|opposite|campus|'
                r'complex|block|sciences|research|govt|government|medical|centre|center|building|'
                r'plot|phase|sector|village|taluk|tehsil|dept|www|@|\d{4,}',re.I)
CITYFIX={'esi co orporation':'Ludhiana','gradi gate':'Bhavnagar','saifai etawah':'Etawah',
 'skims mc bemina':'Srinagar','dhanbad)':'Dhanbad','beltola':'Guwahati','joka':'Kolkata',
 'lehriasarai':'Darbhanga','panaji':'Panaji Goa','bambolim':'Panaji Goa','nanda nagar':'Indore',
 'arrah da':'Patna','kamarhati':'Kolkata','pudukottai':'Pudukkottai','burdwan':'Bardhaman (Burdwan)'}
def cityof(full):
    parts=[p.strip() for p in full.split(',')]
    cand=''
    for p in parts[1:]:
        if p and len(p.split())<=3 and not _BAD.search(p): cand=p; break
    if not cand:
        for p in parts[1:]:
            if p and len(p.split())<=4: cand=p; break
    c=CITYFIX.get(cand.lower().strip(), cand)
    return re.sub(r'\s+',' ',c).strip(' ,-()')
def propername(full):
    parts=[p.strip() for p in full.split(',')]
    nm=parts[0]
    ct=cityof(full)
    if ct and ct.lower() not in nm.lower():
        nm=f'{nm}, {ct}'
    return re.sub(r'\s+',' ',nm).strip(' ,')
# city -> (airport, IATA, km from campus, note).  ESTIMATES from public geography - verify.
AIR={
'indore':('Devi Ahilya Bai Holkar','IDR',10,'In city; LKO direct'),
'raipur':('Swami Vivekananda','RPR',15,'In city; LKO direct'),
'kolkata':('Netaji Subhash Chandra Bose','CCU',16,'RG Kar ~16 km, Sagar Dutta ~15 km, Joka ~35 km; LKO direct'),
'patna':('Jay Prakash Narayan','PAT',10,'In city; LKO direct'),
'bardhaman (burdwan)':('Kolkata CCU','CCU',105,'2.5-3 hr road; LKO direct to Kolkata only'),
'guwahati':('Lokpriya Gopinath Bordoloi','GAU',25,'LKO direct'),
'panaji goa':('Dabolim / Mopa','GOI / GOX',30,'Dabolim ~30 km, Mopa ~35 km; LKO direct'),
'kalyani':('Kolkata CCU','CCU',60,'~1.5 hr road; LKO direct to Kolkata'),
'baramati':('Pune','PNQ',100,'~2.5 hr road; LKO direct to Pune'),
'karur':('Tiruchirappalli','TRZ',80,'Coimbatore CJB ~130 km alt; no LKO direct'),
'namakkal':('Salem','SXV',55,'Trichy TRZ ~95 km alt; no LKO direct'),
'shillong':('Guwahati GAU','GAU',120,'Shillong SHL ~30 km but very limited; LKO direct to Guwahati'),
'dhanbad':('Durgapur / Ranchi','RDP / IXR',90,'Durgapur ~90 km, Ranchi ~140 km; strong direct rail from LKO'),
'mandya':('Mysuru / Bengaluru','MYQ / BLR',45,'Mysuru ~45 km limited, Bengaluru ~130 km; no LKO direct'),
'rewa':('Rewa (new) / Prayagraj','REW / IXD',15,'Rewa airport new and limited; Prayagraj ~230 km'),
'jabalpur':('Dumna','JLR',20,'~64 flights/wk; NO LKO direct - one stop'),
'ludhiana':('Sahnewal / Chandigarh','LUH / IXC',10,'Sahnewal very limited; Chandigarh ~100 km, LKO direct'),
'varanasi':('Lal Bahadur Shastri','VNS',25,'In city; ~320 km by road from Lucknow'),
'hassan':('Mangaluru / Bengaluru','IXE / BLR',170,'Both far; no practical day trip'),
'gulbarga':('Kalaburagi / Hyderabad','GBI / HYD',15,'Kalaburagi limited; Hyderabad ~220 km'),
'idukki':('Cochin','COK',110,'Madurai IXM ~140 km alt'),
'dharmapuri':('Salem / Hosur','SXV',65,'Hosur ~90 km alt'),
'pali':('Jodhpur','JDH',75,'No LKO direct'),
'pudukkottai':('Tiruchirappalli','TRZ',55,'No LKO direct'),
'bhilwara':('Udaipur','UDR',160,'Jaipur ~250 km alt'),
'bhavnagar':('Bhavnagar','BHU',10,'In city; no LKO direct'),
'jammu':('Jammu','IXJ',5,'In city'),
'etawah':('Kanpur / Lucknow','KNU / LKO',130,'Drivable from Lucknow ~230 km - no flight needed'),
'churu':('Jaipur / Bikaner','JAI / BKB',180,'Both far'),
'srinagar':('Sheikh ul-Alam','SXR',12,'In city'),
'nagaur':('Jodhpur','JDH',135,'No LKO direct'),
'faridkot':('Bathinda / Amritsar','BUP / ATQ',60,'Chandigarh ~200 km alt'),
'bharatpur':('Agra / Delhi','AGR / DEL',55,'Delhi ~185 km, LKO direct to Delhi'),
'konni':('Trivandrum','TRV',100,'Cochin ~130 km alt'),
'darbhanga':('Darbhanga','DBR',10,'In city; Patna ~140 km alt'),
'jamshedpur':('Sonari / Ranchi','IXW / IXR',5,'Sonari very limited; Ranchi ~130 km'),
'sonepat':('Delhi','DEL',50,'LKO direct to Delhi'),
'theni':('Madurai','IXM',80,'No LKO direct'),
'bhopal':('Raja Bhoj','BHO',15,'In city; LKO direct'),
'jaipur':('Jaipur','JAI',12,'In city; LKO direct'),
'chennai':('Chennai','MAA',15,'In city; LKO direct'),
'delhi':('Indira Gandhi','DEL',15,'In city; LKO direct - most frequent route'),
}
def airportof(full):
    fl=full.lower()
    # campus-specific overrides inside one metro
    if 'joka' in fl: return ('Netaji Subhash Chandra Bose','CCU',35,'Joka, South Kolkata ~1 hr; LKO direct')
    if 'kamarhati' in fl or 'sagore dutta' in fl: return ('Netaji Subhash Chandra Bose','CCU',15,'Kamarhati ~40 min; LKO direct')
    if 'rg kar' in fl: return ('Netaji Subhash Chandra Bose','CCU',16,'Shyambazar ~40 min; LKO direct')
    c=cityof(full).lower()
    if c in AIR: return AIR[c]
    for k,v in AIR.items():
        if k in c or c in k: return v
    return None

STOP={'government','govt','goverment','medical','college','hospital','institute','of','and','the',
 'sciences','science','dr','memorial','research','centre','center','general','district','state'}
def band(c): return 7000<=c<=13000
out=[]
for c,o,s,inst in rows:
    if not band(c): continue
    low=inst.lower()
    parts=[p.strip().lower() for p in inst.split(',')]
    # match ONLY against college name + city field, never the street address
    cityscope=' '.join(parts[:2])
    d=None
    for k,v in DIRECT.items():
        if re.search(r'\b'+re.escape(k)+r'\b',cityscope): d=v; break
    # per-campus surface leg overrides; None => too far to be a day trip
    OVR={'sagore dutta':('Kolkata (CCU)','~15 km / ~40 min','Direct; comfortable day trip'),
         'rg kar':('Kolkata (CCU)','~16 km / ~40 min','Direct; comfortable day trip'),

         'joka':('Kolkata (CCU)','~35 km / ~1 hr','Direct; workable day trip'),
         'ludhiana':(None,None,None), 'varanasi':(None,None,None), 'gulbarga':(None,None,None),
         'nanda nagar':('Indore (IDR)','~10 km / ~25 min','Direct; best-shaped day trip in the band'),
         'arrah da':('Patna (PAT)','~10 km / ~30 min','Direct; comfortable day trip'),
         'beltola':('Guwahati (GAU)','~25 km / ~45 min','Direct; workable day trip'),
         'kalyani':('Kolkata (CCU)','~60 km / ~1.5 hr','Direct city but a long transfer'),
         'burdwan':('Kolkata (CCU)','~105 km / ~2.5-3 hr','Direct city, but 5-6 hr road round trip on top of flights - marginal'),
         'mandya':(None,None,None),
         'raichur':(None,None,None),
         'baramati':('Pune (PNQ)','~100 km / ~2.5 hr','Direct city but a long transfer'),
         'shillong':(None,None,None),
         'neigrihms':(None,None,None)}
    for k,v in OVR.items():
        if k in low:
            d = v if v[0] else None
            break
    pv=None
    for k,v in PRIV.items():
        if k in low: pv=v; break
    fe=None
    for k,v in FEE.items():
        if k in low: fe=v; break
    out.append({'close':c,'open':o,'seats':s,'name':propername(inst),'city':cityof(inst),
                'air':airportof(inst),'full':inst,
                'direct':d,'priv':pv,'fee':fe})
# rank: P1 direct flight, P2 fee<35k, P3 privacy
def p1(r): return 0 if r['direct'] else 1
def p2(r):
    if r['fee'] and r['fee'][0]<35000: return 0
    if not r['fee']: return 1
    return 2
def p3(r):
    if not r['priv']: return 3
    v=r['priv'][0]
    if v.startswith('CLEAN'): return 0
    if v.startswith('Clean'): return 1
    if 'stale' in v or 'hidden' in v: return 2
    if v.startswith('UNVERIFIED') or v.startswith('Unverified') or v.startswith('Unresolved'): return 2
    return 4
out.sort(key=lambda r:(p1(r),p2(r),p3(r),r['close']))
json.dump(out,open('ranked.json','w'),indent=1,default=str)
print('band rows:',len(out))
print('with direct flight:',sum(1 for r in out if r['direct']))
print('with a privacy finding:',sum(1 for r in out if r['priv']))
print('with a confirmed fee:',sum(1 for r in out if r['fee']))


# ---- merge my earlier all-135 workbook (calendar, fee, documents-found, evidence, state)
import openpyxl as _op
PRIOR={}
try:
    _w=_op.load_workbook('/home/user/sorcery/audit/aiq-privacy/inputs/AIQ_all_colleges_fees_calendar.xlsx')
    _s=_w['All colleges - filters']
    for _r in range(2,_s.max_row+1):
        nm=(_s.cell(_r,1).value or '').strip()
        if not nm: continue
        PRIOR[nm.lower()]={'state':_s.cell(_r,2).value,'c25':_s.cell(_r,4).value,
          'cal':_s.cell(_r,5).value,'calev':_s.cell(_r,6).value,'fee':_s.cell(_r,7).value,
          'basis':_s.cell(_r,9).value,'docs':_s.cell(_r,10).value,'ev':_s.cell(_r,11).value,
          'site':_s.cell(_r,12).value}
    print('merged prior workbook rows:',len(PRIOR))
except Exception as e:
    print('prior workbook unavailable:',e)
def prior_for(name):
    ln=name.lower()
    toks={w for w in re.sub(r'[^a-z ]',' ',ln).split() if w not in STOP and len(w)>3}
    best=None
    for k,v in PRIOR.items():
        kt={w for w in re.sub(r'[^a-z ]',' ',k).split() if w not in STOP and len(w)>3}
        ov=len(toks & kt)
        if ov and (best is None or ov>best[0]): best=(ov,v)
    return best[1] if best and best[0]>=2 else None
for o in out:
    o['prior']=prior_for(o['full'])
    p=o['prior']
    if p:
        if not o['fee'] and isinstance(p.get('fee'),int): o['fee']=(p['fee'], p.get('basis') or 'from prior workbook')
        if not o['priv'] and p.get('docs'): o['priv']=('From prior audit','', p['docs'])
out.sort(key=lambda r:(p1(r),p2(r),p3(r),r['close']))
json.dump(out,open('ranked.json','w'),indent=1,default=str)
# ================= build the workbook =================
wb=openpyxl.Workbook(); ws=wb.active; ws.title='Ranked - flight, fee, privacy'
HDR=Font(bold=True,color='FFFFFF',size=10); HFIL=PatternFill('solid',fgColor='1F4E79')
WRAP=Alignment(wrap_text=True,vertical='top'); CEN=Alignment(horizontal='center',vertical='center')
TH=Border(*[Side(style='thin',color='D9D9D9')]*4)
G=PatternFill('solid',fgColor='C6EFCE'); Y=PatternFill('solid',fgColor='FFEB9C')
R=PatternFill('solid',fgColor='FFC7CE'); GR=PatternFill('solid',fgColor='F2F2F2')
COLS=[('Rank',6),('College',44),('2026 R1 close',11),('Open',9),('R1 Gen seats',9),
      ('P1: Direct LKO flight?',13),('Airport',20),('Airport to campus',18),('Flight note',30),
      ('P2: Fee under Rs 35,000?',13),('Confirmed fee Rs/yr',12),('Fee basis',40),
      ('P3: Privacy verdict',24),('Newest list found',12),('Privacy detail - what was actually found',96),
      ('City',20),('Nearest airport',30),('IATA',12),('Km airport to campus',13),('Airport note',40),
      ('State',16),('2025 R1 close',11),('Academic calendar?',13),('Evidence URL',46),('Site',26)]
for j,(h,w) in enumerate(COLS,1):
    c=ws.cell(1,j,h); c.font=HDR; c.fill=HFIL
    c.alignment=Alignment(wrap_text=True,horizontal='center',vertical='center')
    ws.column_dimensions[get_column_letter(j)].width=w
ws.row_dimensions[1].height=48
data=json.load(open('ranked.json'))
r=2
for i,o in enumerate(data,1):
    d=o['direct']; pv=o['priv']; fe=o['fee']
    fyes='YES' if (fe and fe[0]<35000) else ('UNKNOWN' if not fe else 'NO')
    vals=[i,o['name'],o['close'],o['open'],o['seats'],
          'YES' if d else 'no', d[0] if d else '', d[1] if d else '', d[2] if d else '',
          fyes, (fe[0] if fe else ''), (fe[1] if fe else 'NOT CONFIRMED - no fee document reachable'),
          (pv[0] if pv else 'Not audited'), (pv[1] if pv else ''),
          (pv[2] if pv else 'Not audited - site unreachable from this environment, or never in the worklist'),
          o['city'], (o['air'][0] if o['air'] else 'not resolved'),
          (o['air'][1] if o['air'] else ''), (o['air'][2] if o['air'] else ''),
          (o['air'][3] if o['air'] else ''),
          (o['prior'] or {}).get('state',''), (o['prior'] or {}).get('c25',''),
          (o['prior'] or {}).get('cal',''), (o['prior'] or {}).get('ev',''), (o['prior'] or {}).get('site','')]
    for j,v in enumerate(vals,1):
        c=ws.cell(r,j,v); c.border=TH
        c.alignment=WRAP if j in (9,12,15,20) else CEN
    ws.cell(r,6).fill = G if d else GR
    ws.cell(r,10).fill = G if fyes=='YES' else (GR if fyes=='UNKNOWN' else R)
    pvv=(pv[0] if pv else '')
    ws.cell(r,13).fill = G if pvv.startswith('CLEAN') else (Y if (pvv.startswith(('Clean','UNVERIFIED','Unverified','Unresolved')) or 'stale' in pvv or 'hidden' in pvv) else (R if pvv else GR))
    ws.cell(r,11).number_format='#,##0'
    r+=1
ws.auto_filter.ref=f'A1:Y{r-1}'; ws.freeze_panes='C2'

ns=wb.create_sheet('How this is ranked')
N=[('Priority 1 - Connectivity','Direct non-stop flight from Lucknow (LKO). Confirmed direct destinations as of Aug 2026: '
   'Delhi/Noida, Kolkata, Hyderabad, Indore, Raipur, Patna, Pune, Bengaluru, Mumbai, Goa, Guwahati, Jaipur, '
   'Dehradun, Chandigarh, Chennai, Bhopal, Jharsuguda, Kishangarh(Ajmer). NOT direct: Nagpur, Jabalpur, '
   'Bilaspur, Ranchi, Bhubaneswar. Only 10 of the 139 in-band colleges sit in a directly-connected city.'),
 ('Priority 2 - Fees','Under Rs 35,000/yr. A figure appears ONLY where a college document states it. '
   'Aggregator sites were rejected - they returned Rs 70,500 / 1,62,100 / 6,92,000 for the same Maharashtra '
   'colleges. Only 7 in-band colleges have any document-backed fee at all. NOTE: CIMS Bilaspur at Rs 40,000 '
   'now FAILS this ceiling (it passed the earlier Rs 40,000 one).'),
 ('Priority 3 - Privacy','Green = CLEAN under a full battery. Amber = clean on a lighter pass, or a roster that '
   'is deeply hidden or stale (pre-2020), or unverified. Red = a current, plainly-published list. '
   'Grey = never audited.'),
 ('Why so many blanks','139 colleges fall in the 7,000-13,000 band; I have privacy findings for 33 and '
   'document-backed fees for 7. The rest sit on geo-blocked state portals unreachable from this environment. '
   'A blank is "unknown", never "clean".'),
 ('Known defects in my own data','1) Tiruppur, Namakkal and Ariyalur are recorded with a student-merit-list URL '
   'in the evidence column while the summary column said "no documents found". Treat all three as UNVERIFIED. '
   '2) Single-page-app sites return an identical shell on every path, so a crawl reporting 0 documents proves '
   'nothing - this produced a false clean for Jabalpur until its API was mined. Mandya may have the same flaw. '
   '3) 13 of my original 38 PASSes proved wrong on re-check; every correction came from the user or a later '
   're-run, never from the original method.'),
 ('Not in this sheet','Latur (13,905), Akola (13,994), Ambajogai (14,470), Chandrapur (15,048), Seoni (15,877), '
   'Ratnagiri (16,082), Bhandara (18,616) and Dhanbad-adjacent options above 13,000 are excluded by the band. '
   'Dhanbad itself (10,303) IS in band and is the best-verified clean result.'),
 ('Airport distances','My estimates from public geography, not measured. Verify before booking.'),
 ('MCC 2026 reporting','Under the 2026 rules, FLOAT candidates upload scans and the college verifies them ONLINE '
   'with no visit. Only FREEZE requires the single physical trip. If you float, connectivity stops mattering.')]
for j,(h,w) in enumerate([('Topic',28),('Detail',150)],1):
    c=ns.cell(1,j,h); c.font=HDR; c.fill=HFIL; ns.column_dimensions[get_column_letter(j)].width=w
for i,(a,b) in enumerate(N,2):
    ns.cell(i,1,a).alignment=WRAP; ns.cell(i,2,b).alignment=WRAP

# colour the km column for the top 40
for rr in range(2, min(42, r)):
    km=ws.cell(rr,19).value
    if isinstance(km,int): ws.cell(rr,19).fill = G if km<=35 else (Y if km<=100 else R)
wb.save('AIQ_ranked_flight_fee_privacy.xlsx')
print('saved. top 12:')
for i,o in enumerate(data[:12],1):
    print(f"{i:2d}. {o['close']:>6,} {o['name'][:40]:40s} direct={'Y' if o['direct'] else 'n'} fee={o['fee'][0] if o['fee'] else '?'} priv={(o['priv'][0] if o['priv'] else 'not audited')[:30]}")
